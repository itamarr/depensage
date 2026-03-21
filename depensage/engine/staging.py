"""
Staging layer for the expense pipeline.

Collects all pending writes into MonthStage objects. The caller
can inspect, display, export to XLSX, and then commit (or discard).
"""

import logging
import os
from dataclasses import dataclass, field
from datetime import datetime

from depensage.engine.bank_parser import CCLumpSum

logger = logging.getLogger(__name__)


@dataclass
class RowMeta:
    """Original classification at staging time, for diff detection."""
    orig_category: str
    orig_subcategory: str
    needs_review: bool = False  # True for unclassified rows


@dataclass
class RowChange:
    """A classification change detected in the edited XLSX."""
    month: str
    row_type: str          # "expense" or "income"
    source: str            # "cc", "bank", "income"
    lookup_key: str        # column B (expenses) or column D (income)
    old_category: str
    new_category: str
    old_subcategory: str
    new_subcategory: str


@dataclass
class MonthStage:
    """Staged writes for a single month sheet."""
    month: str
    year: int
    new_expenses: list[list] = field(default_factory=list)  # B:H rows
    new_income: list[list] = field(default_factory=list)    # D:G rows
    expense_meta: list[RowMeta] = field(default_factory=list)  # parallel to new_expenses
    income_meta: list[RowMeta] = field(default_factory=list)   # parallel to new_income
    expense_start_row: int = 0
    income_start_row: int | None = None
    expense_insert_needed: int = 0
    income_insert_needed: int = 0
    duplicates: int = 0
    income_duplicates: int = 0
    bank_balance: float | None = None  # closing bank balance for E175
    savings_allocations: list = field(default_factory=list)  # SavingsAllocation list
    savings_warning: str | None = None
    is_new: bool = False  # True if month sheet needs creation at commit
    carryover_updates: list[tuple] = field(default_factory=list)  # (cell_ref, formula_or_value)


def _month_order(month_name):
    """Return sort key for English month names."""
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12,
    }
    return months.get(month_name, 0)


@dataclass
class StagedPipelineResult:
    """Pipeline output with all writes staged (not yet committed)."""
    # Statistics
    total_parsed: int = 0
    in_process_skipped: int = 0
    classified: int = 0
    unclassified: int = 0
    unclassified_merchants: list[str] = field(default_factory=list)
    cc_lump_sums: list[CCLumpSum] = field(default_factory=list)

    # Staged writes
    month_stages: dict[tuple, MonthStage] = field(default_factory=dict)

    def get_or_create_stage(self, sheet_name, year):
        """Get or create a MonthStage for the given month/year."""
        key = (sheet_name, year)
        if key not in self.month_stages:
            self.month_stages[key] = MonthStage(month=sheet_name, year=year)
        return self.month_stages[key]

    def sorted_stages(self):
        """Return month stages sorted by year then month."""
        return sorted(
            self.month_stages.values(),
            key=lambda s: (s.year, _month_order(s.month)),
        )

    def summary(self):
        """Human-readable summary for CLI stdout."""
        lines = [
            f"  Parsed:     {self.total_parsed}",
            f"  In-process: {self.in_process_skipped} (skipped)",
            f"  Classified: {self.classified}",
            f"  Unknown:    {self.unclassified}",
            "",
        ]

        for stage in self.sorted_stages():
            parts = []
            if stage.new_expenses:
                parts.append(f"{len(stage.new_expenses)} expenses")
            if stage.duplicates:
                parts.append(f"{stage.duplicates} duplicates")
            if stage.new_income:
                parts.append(f"{len(stage.new_income)} income")
            if stage.income_duplicates:
                parts.append(f"{stage.income_duplicates} income dupes")
            if stage.bank_balance is not None:
                parts.append(f"bank balance: {stage.bank_balance:,.2f}")
            if stage.savings_allocations:
                parts.append(f"{len(stage.savings_allocations)} savings allocations")
            if stage.carryover_updates:
                parts.append(f"{len(stage.carryover_updates)} carryover updates")
            if stage.is_new:
                parts.append("NEW sheet")
            if parts:
                lines.append(f"  {stage.month} {stage.year}: {', '.join(parts)}")
            if stage.savings_warning:
                lines.append(f"    Warning: {stage.savings_warning}")

        if self.unclassified_merchants:
            lines.append(f"\n  Unclassified merchants ({len(self.unclassified_merchants)}):")
            for name in self.unclassified_merchants:
                lines.append(f"    - {name}")

        if self.cc_lump_sums:
            total = sum(ls.amount for ls in self.cc_lump_sums)
            lines.append(
                f"\n  CC lump sums from bank ({len(self.cc_lump_sums)}): "
                f"{total:,.2f}"
            )
            for ls in self.cc_lump_sums:
                date_str = (
                    ls.date.strftime("%Y-%m-%d")
                    if hasattr(ls.date, "strftime") else str(ls.date)
                )
                lines.append(f"    - {ls.amount:,.2f} ({date_str})")

        return "\n".join(lines)

    def has_writes(self):
        """Return True if there are any staged writes."""
        return any(
            stage.new_expenses or stage.new_income
            or stage.bank_balance is not None
            or stage.savings_allocations
            or stage.carryover_updates
            for stage in self.month_stages.values()
        )

    def export_xlsx(self, path=None, categories_with_subcats=None):
        """Export staged changes to XLSX for review.

        Args:
            path: Output path. Defaults to .artifacts/staged_<timestamp>.xlsx.
            categories_with_subcats: Optional set of category names that have
                subcategories. Used for red-highlighting: empty subcategory is
                only flagged when the category has defined subcategories.

        Returns:
            File path of the exported XLSX.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        if path is None:
            os.makedirs(".artifacts", exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = f".artifacts/staged_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        # Create summary sheet first
        summary_ws = wb.active
        summary_ws.title = "Summary"

        bold = Font(bold=True)
        summary_ws.append(["Month", "New Expenses", "Duplicates", "New Income", "Income Dupes", "Bank Balance"])
        for cell in summary_ws[1]:
            cell.font = bold

        for stage in self.sorted_stages():
            summary_ws.append([
                f"{stage.month} {stage.year}",
                len(stage.new_expenses),
                stage.duplicates,
                len(stage.new_income),
                stage.income_duplicates,
                stage.bank_balance if stage.bank_balance is not None else "",
            ])

        # Add unclassified merchants to summary
        if self.unclassified_merchants:
            summary_ws.append([])
            summary_ws.append(["Unclassified merchants"])
            summary_ws[summary_ws.max_row][0].font = bold
            for name in self.unclassified_merchants:
                summary_ws.append([name])

        # Per-month sheets with new rows
        expense_headers = ["שם בית עסק", "הערות", "תת קטגוריה", "כמה", "קטגוריה", "תאריך", "סטטוס"]
        income_headers = ["הערות", "כמה", "קטגוריה", "תאריך"]

        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        bank_balance_label = 'יתרה בעו"ש'
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red_warning_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        for stage in self.sorted_stages():
            if (not stage.new_expenses and not stage.new_income
                    and stage.bank_balance is None
                    and not stage.savings_allocations):
                continue

            ws_name = f"{stage.month[:3]} {stage.year}"
            ws = wb.create_sheet(title=ws_name)

            if stage.new_expenses:
                ws.append(expense_headers)
                for cell in ws[1]:
                    cell.font = bold
                for row in stage.new_expenses:
                    ws.append(row)
                # Red-highlight empty category cells, and empty subcategory
                # only when the category has defined subcategories
                has_subcats = categories_with_subcats or set()
                for row_idx in range(2, len(stage.new_expenses) + 2):
                    cat_cell = ws.cell(row=row_idx, column=5)   # column E = category
                    sub_cell = ws.cell(row=row_idx, column=3)   # column C = subcategory
                    if not cat_cell.value:
                        cat_cell.fill = red_fill
                    elif not sub_cell.value and cat_cell.value in has_subcats:
                        sub_cell.fill = red_fill

            if stage.new_income:
                if stage.new_expenses:
                    ws.append([])  # blank separator
                ws.append(income_headers)
                header_row = ws.max_row
                for cell in ws[header_row]:
                    cell.font = bold
                for row in stage.new_income:
                    ws.append(row)

            if stage.bank_balance is not None:
                if stage.new_expenses or stage.new_income:
                    ws.append([])  # blank separator
                ws.append([bank_balance_label, stage.bank_balance])
                bal_row = ws.max_row
                ws.cell(row=bal_row, column=1).font = bold

            if stage.savings_allocations:
                ws.append([])  # blank separator
                ws.append(["Savings Allocation"])
                ws.cell(row=ws.max_row, column=1).font = bold
                ws.append(["Goal", "Preset", "Allocated", "Target", "Current Total"])
                for cell in ws[ws.max_row]:
                    if cell.value:
                        cell.font = bold

                for alloc in stage.savings_allocations:
                    label = alloc.goal_name
                    if alloc.is_default:
                        label += " (default)"
                    ws.append([
                        label,
                        alloc.preset_incoming or "",
                        alloc.allocated,
                        alloc.target or "",
                        alloc.current_total or "",
                    ])

                if stage.savings_warning:
                    ws.append([])
                    ws.append([stage.savings_warning])
                    warn_row = ws.max_row
                    warn_fill = red_warning_fill if any(
                        a.allocated == 0 for a in stage.savings_allocations
                    ) else yellow_fill
                    ws.cell(row=warn_row, column=1).fill = warn_fill

        # Add hidden metadata sheet for edit-back flow
        meta_ws = wb.create_sheet(title="_row_meta")
        meta_ws.sheet_state = "hidden"
        meta_ws.append(["month", "row_type", "row_index", "orig_category", "orig_subcategory"])
        for stage in self.sorted_stages():
            ws_name = f"{stage.month[:3]} {stage.year}"
            for i, meta in enumerate(stage.expense_meta):
                meta_ws.append([ws_name, "expense", i, meta.orig_category, meta.orig_subcategory])
            for i, meta in enumerate(stage.income_meta):
                meta_ws.append([ws_name, "income", i, meta.orig_category, meta.orig_subcategory])
            for i, alloc in enumerate(stage.savings_allocations):
                meta_ws.append([ws_name, "savings", i, alloc.goal_name, alloc.row_number])
            if stage.is_new:
                meta_ws.append([ws_name, "is_new", 0, "", ""])
            for i, (ref, val) in enumerate(stage.carryover_updates):
                # Escape formulas so openpyxl doesn't interpret them.
                # Without this, ='January'!C159 gets mangled to
                # =[1]January!C159 (external workbook reference) because
                # the XLSX file has no sheet named "January".
                if isinstance(val, str) and val.startswith("="):
                    val = "FORMULA:" + val[1:]
                meta_ws.append([ws_name, "carryover", i, ref, val])

        wb.save(path)
        return path

    def commit(self, handlers):
        """Execute all staged writes to the spreadsheet.

        Processes months in chronological order. Batches cell updates
        (savings allocations, bank balance) to minimize API calls.

        Args:
            handlers: Dict mapping year (int) to SheetHandler,
                      or a single SheetHandler.

        Returns:
            Committed PipelineResult with final statistics.
        """
        from depensage.engine.pipeline import MonthResult, PipelineResult

        if not isinstance(handlers, dict):
            single = handlers
            get_handler = lambda y: single
        else:
            def get_handler(y):
                return handlers[y]

        month_results = []
        for stage in self.sorted_stages():
            handler = get_handler(stage.year)
            written = 0
            income_written = 0

            # 1. Create sheet from template if new
            if stage.is_new:
                success = handler.create_sheet_from_template(stage.month)
                if not success:
                    logger.error(
                        f"Failed to create sheet '{stage.month}', skipping"
                    )
                    continue

            # 2. Write carryover formulas/values
            if stage.carryover_updates:
                handler.batch_update_cells(stage.month, stage.carryover_updates)
                logger.info(
                    f"Wrote {len(stage.carryover_updates)} carryover "
                    f"updates to {stage.month}"
                )

            # 3. Write expenses
            if stage.new_expenses:
                if stage.expense_insert_needed > 0:
                    marker = handler.find_section_marker(stage.month, "budget")
                    if marker:
                        handler.insert_rows(
                            stage.month, marker - 1, stage.expense_insert_needed
                        )
                handler.write_expense_rows(
                    stage.month, stage.expense_start_row, stage.new_expenses
                )
                written = len(stage.new_expenses)

            # Write income
            if stage.new_income and stage.income_start_row is not None:
                if stage.income_insert_needed > 0:
                    savings_marker = handler.find_section_marker(
                        stage.month, "savings"
                    )
                    if savings_marker:
                        handler.insert_rows(
                            stage.month, savings_marker - 2,
                            stage.income_insert_needed,
                        )
                handler.write_income_rows(
                    stage.month, stage.income_start_row, stage.new_income
                )
                income_written = len(stage.new_income)

            # Batch cell updates: savings allocations + bank balance
            cell_updates = []

            if stage.savings_allocations:
                from depensage.engine.savings_allocator import find_savings_goal_rows
                fresh_rows = find_savings_goal_rows(handler, stage.month)
                for alloc in stage.savings_allocations:
                    row_num = fresh_rows.get(alloc.goal_name, alloc.row_number)
                    if row_num:
                        cell_updates.append((f"E{row_num}", alloc.allocated))

            if stage.bank_balance is not None:
                bal_row = handler.find_reconciliation_label_row(
                    stage.month, 'כסף בעו"ש'
                )
                if bal_row:
                    cell_updates.append((f"E{bal_row}", stage.bank_balance))
                else:
                    logger.warning(
                        f"Could not find bank balance label in "
                        f"{stage.month} reconciliation section"
                    )

            if cell_updates:
                handler.batch_update_cells(stage.month, cell_updates)

            month_results.append(MonthResult(
                month=stage.month,
                year=stage.year,
                written=written,
                duplicates=stage.duplicates,
                income_written=income_written,
                income_duplicates=stage.income_duplicates,
            ))

        return PipelineResult(
            total_parsed=self.total_parsed,
            in_process_skipped=self.in_process_skipped,
            classified=self.classified,
            unclassified=self.unclassified,
            months=month_results,
            unclassified_merchants=self.unclassified_merchants,
            cc_lump_sums=self.cc_lump_sums,
        )



def _cell_str(value):
    """Normalize a cell value to a string, treating None as empty."""
    if value is None:
        return ""
    return str(value).strip()


def import_staged_xlsx(path):
    """Read an edited XLSX and reconstruct MonthStage data with change detection.

    Returns:
        (stages_dict, changes) where:
        - stages_dict: {month_ws_name: MonthStage} with edited rows
        - changes: list[RowChange] of classification edits
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)

    # Read metadata
    if "_row_meta" not in wb.sheetnames:
        raise ValueError("XLSX missing _row_meta sheet — not a staged export")

    meta_ws = wb["_row_meta"]
    meta_rows = list(meta_ws.iter_rows(min_row=2, values_only=True))
    # Group by (month, row_type)
    meta_by_key = {}
    savings_meta_by_month = {}  # {ws_name: [(goal_name, row_number), ...]}
    new_months = set()  # ws_names that are new (need sheet creation)
    carryover_by_month = {}  # {ws_name: [(cell_ref, value), ...]}
    for row in meta_rows:
        month, row_type, row_index, orig_cat, orig_sub = row
        if row_type == "is_new":
            new_months.add(month)
        elif row_type == "carryover":
            # orig_cat = cell_ref, orig_sub = value (float or formula string)
            ref = _cell_str(orig_cat)
            val = orig_sub
            # Unescape formulas that were escaped during export
            if isinstance(val, str) and val.startswith("FORMULA:"):
                val = "=" + val[8:]
            elif val is not None:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            carryover_by_month.setdefault(month, []).append((ref, val))
        elif row_type == "savings":
            savings_meta_by_month.setdefault(month, []).append(
                (_cell_str(orig_cat), orig_sub)  # goal_name, row_number
            )
        else:
            meta_by_key.setdefault((month, row_type), []).append(
                RowMeta(
                    orig_category=_cell_str(orig_cat),
                    orig_subcategory=_cell_str(orig_sub),
                )
            )

    stages = {}
    changes = []

    for ws_name in wb.sheetnames:
        if ws_name in ("Summary", "_row_meta"):
            continue

        ws = wb[ws_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        # Parse the sheet: find expense header, expense rows, income header, income rows
        expense_headers = (
            "שם בית עסק", "הערות", "תת קטגוריה", "כמה", "קטגוריה", "תאריך", "סטטוס"
        )
        income_headers = ("הערות", "כמה", "קטגוריה", "תאריך")

        # Derive month name and year from ws_name (e.g., "Feb 2026")
        month_abbrevs = {
            "Jan": "January", "Feb": "February", "Mar": "March",
            "Apr": "April", "May": "May", "Jun": "June",
            "Jul": "July", "Aug": "August", "Sep": "September",
            "Oct": "October", "Nov": "November", "Dec": "December",
        }
        parts = ws_name.split()
        month_name = month_abbrevs.get(parts[0], parts[0])
        year = int(parts[1]) if len(parts) > 1 else 0

        stage = MonthStage(
            month=month_name, year=year, is_new=(ws_name in new_months),
            carryover_updates=carryover_by_month.get(ws_name, []),
        )

        i = 0
        # Find and read expenses
        if i < len(rows) and rows[i] and rows[i][0] == expense_headers[0]:
            i += 1  # skip header
            while i < len(rows) and rows[i] and rows[i][0] is not None:
                # Check if this is the income header
                if len(rows[i]) >= 4 and rows[i][0] == income_headers[0]:
                    break
                row = list(rows[i])
                # Ensure 7 columns
                while len(row) < 7:
                    row.append("")
                stage.new_expenses.append([_cell_str(v) for v in row])
                i += 1

        # Skip blank rows
        while i < len(rows) and (not rows[i] or rows[i][0] is None):
            i += 1

        # Find and read income (4 columns only: comments, amount, category, date)
        if i < len(rows) and rows[i] and rows[i][0] == income_headers[0]:
            i += 1  # skip header
            while i < len(rows):
                if not rows[i] or rows[i][0] is None:
                    break
                row = list(rows[i])[:4]  # income is 4 columns
                while len(row) < 4:
                    row.append("")
                stage.new_income.append([_cell_str(v) for v in row])
                i += 1

        # Skip blank rows and look for bank balance label
        while i < len(rows) and (not rows[i] or rows[i][0] is None):
            i += 1
        bank_balance_label = 'יתרה בעו"ש'
        if i < len(rows) and rows[i] and rows[i][0] == bank_balance_label:
            bal_val = rows[i][1]
            if bal_val is not None:
                try:
                    stage.bank_balance = float(bal_val)
                except (ValueError, TypeError):
                    pass
            i += 1

        # Skip blank rows and look for savings allocation header
        while i < len(rows) and (not rows[i] or rows[i][0] is None):
            i += 1
        savings_header_label = "Savings Allocation"
        if i < len(rows) and rows[i] and rows[i][0] == savings_header_label:
            i += 1  # skip header
            # Skip column headers row
            if i < len(rows) and rows[i] and rows[i][0] == "Goal":
                i += 1
            # Read allocation rows
            savings_meta = savings_meta_by_month.get(ws_name, [])
            alloc_idx = 0
            from depensage.engine.savings_allocator import SavingsAllocation
            while i < len(rows) and rows[i] and rows[i][0] is not None:
                # Stop at warning row (no numeric allocation column)
                if rows[i][2] is None:
                    break
                goal_label = str(rows[i][0]).strip()
                # Strip "(default)" suffix
                is_default = "(default)" in goal_label
                goal_name = goal_label.replace(" (default)", "").strip()
                try:
                    allocated = float(rows[i][2]) if rows[i][2] is not None else 0.0
                except (ValueError, TypeError):
                    allocated = 0.0
                # Read context columns
                try:
                    preset = float(rows[i][1]) if rows[i][1] else 0.0
                except (ValueError, TypeError):
                    preset = 0.0
                try:
                    target = float(rows[i][3]) if len(rows[i]) > 3 and rows[i][3] else 0.0
                except (ValueError, TypeError):
                    target = 0.0
                try:
                    current_total = float(rows[i][4]) if len(rows[i]) > 4 and rows[i][4] else 0.0
                except (ValueError, TypeError):
                    current_total = 0.0
                # Get row_number from metadata
                row_number = 0
                if alloc_idx < len(savings_meta):
                    _, row_number = savings_meta[alloc_idx]
                    if row_number is not None:
                        row_number = int(row_number)
                    else:
                        row_number = 0
                stage.savings_allocations.append(SavingsAllocation(
                    goal_name=goal_name,
                    allocated=allocated,
                    row_number=row_number,
                    is_default=is_default,
                    is_blatam=('בלת"ם' in goal_name),
                    preset_incoming=preset,
                    target=target,
                    current_total=current_total,
                ))
                alloc_idx += 1
                i += 1

        stages[ws_name] = stage

        # Detect changes against metadata
        expense_meta = meta_by_key.get((ws_name, "expense"), [])
        for idx, exp_row in enumerate(stage.new_expenses):
            if idx >= len(expense_meta):
                break
            meta = expense_meta[idx]
            new_cat = exp_row[4]   # category column (index 4)
            new_sub = exp_row[2]   # subcategory column (index 2)
            if new_cat != meta.orig_category or new_sub != meta.orig_subcategory:
                # Determine source from status column (index 6)
                status = exp_row[6]
                if status == "BANK":
                    source = "bank"
                else:
                    source = "cc"
                changes.append(RowChange(
                    month=month_name,
                    row_type="expense",
                    source=source,
                    lookup_key=exp_row[0],  # column B
                    old_category=meta.orig_category,
                    new_category=new_cat,
                    old_subcategory=meta.orig_subcategory,
                    new_subcategory=new_sub,
                ))

        income_meta = meta_by_key.get((ws_name, "income"), [])
        for idx, inc_row in enumerate(stage.new_income):
            if idx >= len(income_meta):
                break
            meta = income_meta[idx]
            new_cat = inc_row[2]   # category column (index 2)
            new_sub = inc_row[0]   # comments/subcategory column (index 0)
            if new_cat != meta.orig_category or new_sub != meta.orig_subcategory:
                changes.append(RowChange(
                    month=month_name,
                    row_type="income",
                    source="income",
                    lookup_key=inc_row[0],  # column D
                    old_category=meta.orig_category,
                    new_category=new_cat,
                    old_subcategory=meta.orig_subcategory,
                    new_subcategory=new_sub,
                ))

    total_exp = sum(len(s.new_expenses) for s in stages.values())
    total_inc = sum(len(s.new_income) for s in stages.values())
    logger.info(
        f"Imported {len(stages)} month sheets: "
        f"{total_exp} expenses, {total_inc} income, "
        f"{len(changes)} changes detected"
    )

    return stages, changes
