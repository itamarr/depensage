"""Unit tests for staging XLSX import and change detection."""

import os
import tempfile
import unittest

from depensage.engine.staging import (
    StagedPipelineResult, MonthStage, RowMeta, RowChange,
    import_staged_xlsx,
)


class TestExportImportRoundTrip(unittest.TestCase):
    """Test that export → import preserves data and detects changes."""

    def _make_staged_with_meta(self):
        staged = StagedPipelineResult(total_parsed=3, classified=2, unclassified=1)
        stage = staged.get_or_create_stage("February", 2026)
        stage.new_expenses = [
            ["Shop A", "", "", "100.50", "סופר", "02/01/2026", "CC"],
            ["קופת חולים", "", "קופת חולים", "300.00", "בריאות", "02/05/2026", "BANK"],
            ["Unknown", "", "", "50.00", "", "02/10/2026", "CC"],
        ]
        stage.expense_meta = [
            RowMeta(orig_category="סופר", orig_subcategory=""),
            RowMeta(orig_category="בריאות", orig_subcategory="קופת חולים"),
            RowMeta(orig_category="", orig_subcategory="", needs_review=True),
        ]
        stage.new_income = [
            ["מעסיק", "25000.00", "משכורת", "02/01/2026"],
        ]
        stage.income_meta = [
            RowMeta(orig_category="משכורת", orig_subcategory="מעסיק"),
        ]
        return staged

    def test_round_trip_no_changes(self):
        staged = self._make_staged_with_meta()
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        stages, changes = import_staged_xlsx(path)
        self.assertEqual(len(changes), 0)
        self.assertIn("Feb 2026", stages)
        stage = stages["Feb 2026"]
        self.assertEqual(len(stage.new_expenses), 3)
        self.assertEqual(len(stage.new_income), 1)
        self.assertEqual(stage.month, "February")
        self.assertEqual(stage.year, 2026)

    def test_detect_expense_category_change(self):
        staged = self._make_staged_with_meta()
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        # Edit the XLSX: fill in the unknown row's category
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Feb 2026"]
        # Row 4 is the 3rd data row (1=header, 2-4=data)
        ws.cell(row=4, column=5).value = "שונות"       # category
        ws.cell(row=4, column=3).value = "תת קטגוריה"  # subcategory
        wb.save(path)

        stages, changes = import_staged_xlsx(path)
        self.assertEqual(len(changes), 1)
        c = changes[0]
        self.assertEqual(c.source, "cc")
        self.assertEqual(c.lookup_key, "Unknown")
        self.assertEqual(c.old_category, "")
        self.assertEqual(c.new_category, "שונות")
        self.assertEqual(c.new_subcategory, "תת קטגוריה")

    def test_detect_bank_expense_change(self):
        staged = self._make_staged_with_meta()
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Feb 2026"]
        # Row 3 is the bank row (BANK status)
        ws.cell(row=3, column=5).value = "חשבונות"  # change category
        wb.save(path)

        stages, changes = import_staged_xlsx(path)
        self.assertEqual(len(changes), 1)
        c = changes[0]
        self.assertEqual(c.source, "bank")
        self.assertEqual(c.lookup_key, "קופת חולים")
        self.assertEqual(c.old_category, "בריאות")
        self.assertEqual(c.new_category, "חשבונות")

    def test_detect_income_change(self):
        staged = self._make_staged_with_meta()
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Feb 2026"]
        # Income section: header is after expenses + blank row
        # Expenses: rows 1(header)+2+3+4 = 4 rows, blank row 5, income header row 6
        # Income data starts at row 7
        ws.cell(row=7, column=3).value = "קצבה"  # change income category
        wb.save(path)

        stages, changes = import_staged_xlsx(path)
        self.assertEqual(len(changes), 1)
        c = changes[0]
        self.assertEqual(c.source, "income")
        self.assertEqual(c.old_category, "משכורת")
        self.assertEqual(c.new_category, "קצבה")

    def test_missing_meta_sheet_raises(self):
        import openpyxl
        path = os.path.join(tempfile.mkdtemp(), "bad.xlsx")
        wb = openpyxl.Workbook()
        wb.save(path)

        with self.assertRaises(ValueError) as ctx:
            import_staged_xlsx(path)
        self.assertIn("_row_meta", str(ctx.exception))


    def test_bank_balance_round_trip(self):
        """Bank balance survives export → import round-trip."""
        staged = self._make_staged_with_meta()
        stage = staged.month_stages[("February", 2026)]
        stage.bank_balance = 12345.67

        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        # Verify it's in the XLSX
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["Feb 2026"]
        # Find the balance label row
        found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'יתרה בעו"ש':
                self.assertAlmostEqual(row[1], 12345.67)
                found = True
                break
        self.assertTrue(found, "Bank balance label not found in XLSX")

        # Import and check
        stages, changes = import_staged_xlsx(path)
        self.assertAlmostEqual(stages["Feb 2026"].bank_balance, 12345.67)

    def test_bank_balance_only_stage(self):
        """A month with only bank balance (no expenses/income) is exported."""
        staged = StagedPipelineResult()
        stage = staged.get_or_create_stage("March", 2026)
        stage.bank_balance = 5000.00

        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        import openpyxl
        wb = openpyxl.load_workbook(path)
        self.assertIn("Mar 2026", wb.sheetnames)

        stages, changes = import_staged_xlsx(path)
        self.assertAlmostEqual(stages["Mar 2026"].bank_balance, 5000.00)
        self.assertEqual(len(stages["Mar 2026"].new_expenses), 0)
        self.assertEqual(len(stages["Mar 2026"].new_income), 0)

    def test_no_bank_balance_stays_none(self):
        """Months without bank balance keep bank_balance as None."""
        staged = self._make_staged_with_meta()
        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        stages, changes = import_staged_xlsx(path)
        self.assertIsNone(stages["Feb 2026"].bank_balance)


class TestExportHighlighting(unittest.TestCase):

    def test_empty_category_highlighted_red(self):
        import openpyxl

        staged = StagedPipelineResult()
        stage = staged.get_or_create_stage("January", 2026)
        stage.new_expenses = [
            ["Shop A", "", "", "100.50", "סופר", "01/01/2026", "CC"],
            ["Unknown", "", "", "50.00", "", "01/05/2026", "CC"],
        ]
        stage.expense_meta = [
            RowMeta(orig_category="סופר", orig_subcategory=""),
            RowMeta(orig_category="", orig_subcategory="", needs_review=True),
        ]

        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Jan 2026"]
        # Row 3 (2nd expense, unknown) should have red fill on category (col 5)
        cat_fill = ws.cell(row=3, column=5).fill
        self.assertEqual(cat_fill.start_color.rgb, "00FFCCCC")

        # Row 2 (1st expense, classified) category should NOT be red
        cat_fill_ok = ws.cell(row=2, column=5).fill
        self.assertNotEqual(cat_fill_ok.start_color.rgb, "00FFCCCC")

    def test_empty_subcat_red_only_when_category_has_subcats(self):
        import openpyxl

        staged = StagedPipelineResult()
        stage = staged.get_or_create_stage("January", 2026)
        stage.new_expenses = [
            ["Shop A", "", "", "100.50", "נסיעות", "01/01/2026", "CC"],  # has subcats
            ["Shop B", "", "", "50.00", "אלוני", "01/05/2026", "CC"],    # no subcats
        ]
        stage.expense_meta = [
            RowMeta(orig_category="נסיעות", orig_subcategory=""),
            RowMeta(orig_category="אלוני", orig_subcategory=""),
        ]

        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        # נסיעות has subcats, אלוני does not
        staged.export_xlsx(path, categories_with_subcats={"נסיעות"})

        wb = openpyxl.load_workbook(path)
        ws = wb["Jan 2026"]
        # Row 2: נסיעות with no subcat → subcat should be red
        sub_fill = ws.cell(row=2, column=3).fill
        self.assertEqual(sub_fill.start_color.rgb, "00FFCCCC")

        # Row 3: אלוני with no subcat → subcat should NOT be red
        sub_fill_no = ws.cell(row=3, column=3).fill
        self.assertNotEqual(sub_fill_no.start_color.rgb, "00FFCCCC")


class TestRowMetaInMonthStage(unittest.TestCase):

    def test_meta_exported_to_hidden_sheet(self):
        import openpyxl

        staged = StagedPipelineResult()
        stage = staged.get_or_create_stage("March", 2026)
        stage.new_expenses = [
            ["Shop", "", "sub", "10.00", "cat", "03/01/2026", "CC"],
        ]
        stage.expense_meta = [
            RowMeta(orig_category="cat", orig_subcategory="sub"),
        ]

        path = os.path.join(tempfile.mkdtemp(), "test.xlsx")
        staged.export_xlsx(path)

        wb = openpyxl.load_workbook(path)
        self.assertIn("_row_meta", wb.sheetnames)
        meta_ws = wb["_row_meta"]
        rows = list(meta_ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "Mar 2026")
        self.assertEqual(rows[0][1], "expense")
        self.assertEqual(rows[0][3], "cat")
        self.assertEqual(rows[0][4], "sub")


if __name__ == "__main__":
    unittest.main()
